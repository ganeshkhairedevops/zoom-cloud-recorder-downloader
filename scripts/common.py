# common.py - shared helpers for Zoom recordings downloader
import requests
import os
import openpyxl
from datetime import datetime

# === SAMPLE PLACEHOLDERS (replace locally before running) ===
ACCOUNT_ID = "SAMPLE_ACCOUNT_ID"
CLIENT_ID = "SAMPLE_CLIENT_ID"
CLIENT_SECRET = "SAMPLE_CLIENT_SECRET"
USER_EMAIL = "user@example.com"

DOWNLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "zoom_recordings_one_user")
EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "recordings_report.xlsx")

def get_access_token():
    """Obtain access token using Zoom Server-to-Server OAuth (account_credentials)."""
    url = f"https://zoom.us/oauth/token?grant_type=account_credentials&account_id={ACCOUNT_ID}"
    resp = requests.post(url, auth=(CLIENT_ID, CLIENT_SECRET))
    resp.raise_for_status()
    return resp.json().get("access_token")

def ensure_excel():
    """Create Excel file with header if not exists."""
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recordings"
        ws.append(["File Name", "Recording Date", "Local Path", "Size (MB)"])
        wb.save(EXCEL_PATH)
    return EXCEL_PATH

def save_row_to_excel(row):
    """Append a row to the master excel. Row should be a list."""
    ensure_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.append(row)
    wb.save(EXCEL_PATH)
